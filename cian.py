"""
Программа собирает данные с ЦИАН по заданным ЖК и проводит аналитику.
На вход - список ЖК из таблицы id.xlsx и прокси, на выходе таблица Аналитика на <today>.xlsx
id.xlsx должна быть в директории проекта, Аналитика на <today>.xlsx сохраняется в директорию проекта.
"""


import requests
import os
from bs4 import BeautifulSoup
import re
import numpy as np
import datetime
from fake_useragent import UserAgent
import random
import pandas as pd
from time import sleep
from openpyxl import load_workbook
import pickle
maxpage = 80
random.seed(datetime.datetime.now())
name = 0


def save_cookies(requests_cookiejar, filename):
    with open(filename, 'wb') as f:
        pickle.dump(requests_cookiejar, f)


def load_cookies(filename):
    with open(filename, 'rb') as f:
        return pickle.load(f)


def append_df_to_excel(filename, data_frame, sheet_name='Sheet1', startrow=None,
                       **to_excel_kwargs):
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    # create a writer for this month and year
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if not startrow and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        # copy existing sheets
        writer.sheets = dict(
            (ws.title, ws) for ws in writer.book.worksheets)
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if not startrow:
        startrow = 0

    # write out the new sheet
    data_frame.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def load_data_agent_and_developer_proxy(id, page, proxy):
    url = 'https://www.cian.ru/cat.php?deal_type=sale&engine_version=2&from_developer=2&newobject[0]=%d&offer_type=flat&p=%d'\
        % (id, page)
    if os.path.exists('cookies.pkl'):
        cookie = load_cookies('cookies.pkl')
    else:
        r = requests.get('https://cian.ru/', headers={'User-Agent': UserAgent().random}, proxies={'http': proxy}, cookies={'tmr_detect':'1%7C1554728133459','session_region_id':'175051',
                                                                                                                           'session_main_town_region_id':'175051','cto_lwid':'15116edd-c5eb-47be-85c5-6613ea1b606d',
                                                                                                                           'cto_idcpy':'93918b48-16e9-4ea5-9de7-c08f588518ab',
                                                                                                                           'cfids140':'99XAB/Iei85la2CjKANQFNZGsJ7WQZTXW8t/nFNNYNth88EbrXaLkeBBPy8O6S5igeuAh+fppEPmvRUZpWezSnA7Z6MOOskLPC6zqgYR9+BIPd6pFVPQMau5eAYrVGKfwFvxbe09bQdrvOUkCOzRPIDZHeHOpJs9/JEkUJ0='})
        #cookie = r.cookies
        #save_cookies(r.cookies, 'cookies.pkl')
    #r = requests.get(url, headers={'User-Agent': UserAgent().random}, proxies={'http': proxy}, cookies=cookie)
    return r.text


def load_data_developer_proxy(id, page, proxy):
    url = \
        'https://www.cian.ru/cat.php?deal_type=sale&engine_version=2&from_developer=1&newobject[0]=%d&offer_type=flat&p=%d'\
        % (id, page)
    #cookie = {'tmr_detect':'1%7C1551366696330','session_region_id':'5953','session_main_town_region_id':'175231','cto_lwid':'15116edd-c5eb-47be-85c5-6613ea1b606d','cto_idcpy':'2e25e6da-20ff-41e8-9574-c11bbe084ab6',
              #'cfids140':'WvFwhg+aeZM1Ll7P7ap8v9q09GU/XE/m9ud67AhzUkomVAceKq19HMzDAf9YRQzApRR/UeTARxO1F68PLpBU8ZALQXE1DUprk8D0zsYFmrL9lIAYGdRN8ZZnkw6CUbQCH1n4O2oc8ml2YmTEHK5IbuFD6KT9mxL0rjVTUAPp1P0='}
    if os.path.exists('cookies.pkl'):
        cookie = load_cookies('cookies.pkl')
    else:
        r = requests.get('https://cian.ru/', headers={'User-Agent': UserAgent().random}, proxies={'http': proxy})
        cookie = r.cookies
        save_cookies(r.cookies, 'cookies.pkl')
    r = requests.get(url, headers={'User-Agent': UserAgent().random}, proxies={'http': proxy}, cookies=cookie)
    return r.text


def cian(text):
    global maxpage
    global name
    soup = BeautifulSoup(text, 'lxml')
    price, page, title, flat_type, flat_sqr, floor, id, data, urls, names, dd = \
        [], [], [], [], [], [], [], [], [], [], []
    div_price = soup.find_all('div', class_=re.compile('c6e8ba5398--header'))
    div_max_page = soup.find_all('a', class_=re.compile('_93444fe79c--list'))
    div_title = soup.find_all('div', class_=re.compile('c6e8ba5398--container--F3yyv'))
    div_room_sqr_floor_top = soup.find_all('div', class_=re.compile('c6e8ba5398--single'))
    div_room_sqr_floor = soup.find_all('div', class_=re.compile('c6e8ba5398--title--2CW78'))
    links = soup.find_all('a', class_=re.compile('c6e8ba5398--header'))
    for i in links:
        urls.append(i.attrs['href'])
    name = soup.find('div', class_=re.compile('_93444fe79c--content-title')).a.get_text()
    div_name = soup.find_all('a', class_='c6e8ba5398--building-link--1dQyE')
    div_deadline = soup.find_all('div', class_='c6e8ba5398--deadline--3mUGe')
    #zastroyshik = soup.find('a', class_='_93444fe79c-name--1iqIl').get_text()
    try:
        zastroyshik = soup.find('div', class_=re.compile('c6e8ba5398--name')).get_text()
    except:
        zastroyshik = 'None'
    for i in div_price:
        price.append(re.search('[0-9]*[.,]?[0-9]?[0-9]', re.sub('\s', '', i.get_text())).group(0))
    for i in div_name:
        names.append(i.get_text())
    if div_room_sqr_floor_top != [] and div_room_sqr_floor_top is not None:
        for i in div_room_sqr_floor_top:
            a = re.split(',', i.text)
            flat_sqr.append(re.sub('\D', '', a[1]))
            flat_type.append(re.sub('\D', '', a[0].replace('Студия', '0')))
            floor.append(re.sub('\s', '', a[2].replace('этаж', '')))
            id.append(re.sub('\D', '', i.parent.get('href')))
    if div_room_sqr_floor != [] and div_room_sqr_floor is not None:
        for i in div_room_sqr_floor:
            a = re.split(',', i.text)
            flat_sqr.append(re.sub('\D', '', a[1]))
            flat_type.append(re.sub('\D', '', a[0].replace('Студия', '0')))
            floor.append(re.sub('\s', '', a[2].replace('этаж', '')))
            id.append(re.sub('\D', '', i.parent.get('href')))
    else:
        for i in range(len(price)-3):
            flat_sqr.append('')
            flat_type.append('')
            floor.append('')
            id.append('')
    for i in div_max_page:
        if i.text.isdigit():
            page.append(int(i.get_text()))
    try:
        maxpage = max(page)
    except:
        maxpage = 1
    for i in div_title:
        title.append(re.sub('\n\n*', '', i.get_text().replace('... Подробнее', '')))
    if div_deadline != []:
        for i in div_deadline:
            dd.append(i.get_text().replace('Сдача ГК: ', ''))
    else:
        dd = ['' for i in range(len(price))]
    for i in range(len(price)):
        try:
            if names[i] == name:
                try:
                    data.append({
                        'ID': id[i],
                        'Комнат': int(flat_type[i].replace('-комн. апарт.','')),
                        'Площадь': round(float(flat_sqr[i]),2),
                        'Стоимость': int(price[i]),
                        'Цена за метр': round((float(price[i])/float(flat_sqr[i])),2) - otdelka(zastroyshik),
                        'Этаж': floor[i],
                        'ЖК': name,
                        'Срок сдачи': dd[i],
                        'Застройщик': zastroyshik,
                        'Описание': title[i],
                        'Ссылка': urls[i]
                    })
                except:
                    print('Banned')
                    continue
        except:
            print('Отбрасываем неподходящие квартиры')
            continue
    print(data)
    return data


def otdelka(zastr):
    df = pd.read_excel('otdelka.xlsx')
    try:
        #val = df[df['Застройщик'] == zastr].iloc[0]['Цена']
        val = df[df['ЖК'] == name].iloc[0]['Цена']
        return val
    except IndexError:
        return 0

if __name__ == '__main__':
    list_of_zhk = []
    today = datetime.date.today().strftime("%d-%m-%Y")
    global proxy
    while True:
        try:
            print('Выберите регион мониторинга: 1 - Люберецкий (Новокрасково), 2 - Ленинский, 3 - Произвольный (из таблицы id.xlsx)')
            var = int(input())
            if var == 1:
                list_of_zhk = [6180, 34597, 6233, 39537, 11712, 7484, 7182, 48687, 49046, 49135, 49182, 8160, 51271]  # Люберецкий район
                #list_of_zhk = [49182, 8160, 51271]  # Люберецкий район
                break
            elif var == 2:
                #list_of_zhk = [16751, 19520, 7990, 5198]  # Ленинский район
                list_of_zhk = [7990, 5198]  # Ленинский район
                break
            elif var == 3:
                df = pd.read_excel('id.xlsx')
                for i in range(len(df)):
                    list_of_zhk.append(df.iloc[i]['ID'])
                break
            else:
                print('Неверное знанчение, выберите 1, 2, или 3')
        except ValueError:
            print('Неверное знанчение, выберите 1, 2, или 3')
    print('OK')
    while True:
        try:
            print('Введите прокси в формате IP:порт')
            proxy = input()
            if ':' in proxy:
                break
            else:
                print('Неверное знанчение. Проверьте формат')
        except ValueError:
            print('Неверное знанчение. Проверьте формат')
    print('OK')
    if os.path.exists('cookies.pkl'):
        while(True):
            try:
                print('Удалить старые cookies? Y/N')
                answer = input()
                if answer.upper() == 'Y':
                    os.remove('cookies.pkl')
                    break
                elif answer.upper() == 'N':
                    break
            except:
                print('Неверный ответ')
    for j in range(len(list_of_zhk)):
        data = []
        a = []
        while 1:
            try:
                i = 1
                while (i <= maxpage):
                    text = load_data_developer_proxy(list_of_zhk[j], i, proxy)
                    sp = BeautifulSoup(text, 'lxml')
                    if sp.title.text == 'Captcha - база объявлений ЦИАН':
                        print('CAPTCHA. Необходимо сменить proxy и cookies')
                        break
                    data.append(pd.DataFrame(cian(text)))
                    if maxpage > i:
                        print('Собрано', i, 'страница(ы) из', maxpage)
                    else:
                        print('Собрано', i, 'страница(ы) из', maxpage + 1)
                    i += 1
                    sleep(random.randint(20, 30))
            except Exception as e:
                print(e)
                continue
            break
        df_res = pd.concat(data, ignore_index=True)
        if maxpage > 1:
            df_res.drop_duplicates(subset=['ID'], inplace=True)
        print(df_res)
        for i in range(5):
            try:
                df_1 = df_res[df_res['Комнат'] == i]
                a.append({
                    'Название ЖК': name,
                    'Комнат': i,
                    'Максимальная площадь': df_1['Площадь'].max(axis=0),
                    'Средняя площадь': (df_1['Площадь'].mean(axis=0)).round(2),
                    'Минимальная площадь': df_1['Площадь'].min(axis=0),
                    'Максимальная стоимость': df_1['Стоимость'].max(axis=0),
                    'Средняя стоимость': (df_1['Стоимость'].mean(axis=0)).round(2),
                    'Минимальная стоимость': df_1['Стоимость'].min(axis=0),
                    'Максимальная цена за метр': (df_1['Цена за метр'].max(axis=0)).round(2),
                    'Средняя взвешенная цена за метр': (np.average(df_1['Цена за метр'], weights=df_1['Площадь'])).round(2),
                    'Минимальная цена за метр': (df_1['Цена за метр'].min(axis=0)).round(2),
                    'Объем выборки': df_1['Стоимость'].count()
                })
            except AttributeError:
                print('Attribute Error')
                continue
        analytics = pd.DataFrame(a)
        analytics = analytics[['Название ЖК', 'Комнат', 'Минимальная площадь', 'Средняя площадь',
                               'Максимальная площадь', 'Минимальная цена за метр',
                               'Средняя взвешенная цена за метр', 'Максимальная цена за метр',
                               'Минимальная стоимость', 'Средняя стоимость', 'Максимальная стоимость', 'Объем выборки']]
        print(analytics)
        if os.path.exists('Аналитика на ' + datetime.date.today().strftime("%d-%m-%Y") + '.xlsx'):
            append_df_to_excel('Аналитика на ' + today + '.xlsx', df_res, header=False,
                               sheet_name='Проекты', index=False)
            append_df_to_excel('Аналитика на ' + today + '.xlsx', analytics,
                               header=False, sheet_name='Аналитика', index=False)
        else:
            append_df_to_excel('Аналитика на ' + today + '.xlsx', df_res, sheet_name='Проекты', index=False)
            append_df_to_excel('Аналитика на ' + today + '.xlsx', analytics, sheet_name='Аналитика', index=False)